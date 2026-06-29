"""Build Vercel-ready FedEx IP/IE data from processed IP JSON and IE Excel."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = Path("/Users/alex./Desktop/运费核价助手_FedEx_IP_IE_V1_2_合并计算版_最终修复可打开版.xlsx")
IP_DATA = ROOT / "data_processed" / "fedex_ip_data.json"
OUTPUT = ROOT / "vercel_app" / "data" / "fedex_ip_ie_data.json"


def load_wide_fixed(ws, service: str, source_pages: str) -> list[dict]:
    headers = [cell.value for cell in ws[1]]
    zones = [(idx, str(value)) for idx, value in enumerate(headers) if idx > 0 and value not in (None, "")]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        weight = row[0]
        if weight in (None, ""):
            continue
        for col_idx, zone in zones:
            value = row[col_idx]
            if value in (None, ""):
                continue
            rows.append(
                {
                    "weight_kg": float(weight),
                    "zone": zone,
                    "base_rate_cny": float(value),
                    "source_pdf_pages": source_pages,
                    "service": service,
                }
            )
    return rows


def load_wide_perkg(ws, service: str, source_pages: str) -> list[dict]:
    headers = [cell.value for cell in ws[1]]
    min_col = headers.index("Weight_Min_KG")
    max_col = headers.index("Weight_Max_KG")
    zones = [(idx, str(value)) for idx, value in enumerate(headers) if idx > max_col and value not in (None, "")]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        min_kg = row[min_col]
        max_kg = row[max_col]
        if min_kg in (None, "") or max_kg in (None, ""):
            continue
        for col_idx, zone in zones:
            value = row[col_idx]
            if value in (None, ""):
                continue
            rows.append(
                {
                    "min_kg": float(min_kg),
                    "max_kg": float(max_kg),
                    "zone": zone,
                    "rate_cny_per_kg": float(value),
                    "source_pdf_pages": source_pages,
                    "service": service,
                }
            )
    return rows


def load_validation(ws) -> list[dict]:
    headers = [cell.value for cell in ws[1]]
    valid_headers = [h for h in headers if h not in (None, "")]
    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or raw[0] in (None, ""):
            continue
        row = {key: raw[idx] for idx, key in enumerate(valid_headers)}
        rows.append(row)
    return rows


def main() -> None:
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_EXCEL

    ip_data = json.loads(IP_DATA.read_text(encoding="utf-8"))
    existing_payload = json.loads(OUTPUT.read_text(encoding="utf-8")) if OUTPUT.exists() else {}
    if excel_path.exists():
        wb = load_workbook(excel_path, read_only=False, data_only=True)
        ie_rates = {
            "label": "FedEx IE 国际经济",
            "fixed_0_20_5kg": load_wide_fixed(wb["ie_fixed_wide"], "IE export parcel", "10-12"),
            "perkg_21kg_plus": load_wide_perkg(wb["ie_perkg_wide"], "IE export parcel", "12-13"),
        }
        validation_checks = load_validation(wb["validation_checks"])
        ie_source_excel = excel_path.name
    elif existing_payload.get("rates", {}).get("IE"):
        ie_rates = existing_payload["rates"]["IE"]
        validation_checks = existing_payload.get("validation_checks", [])
        ie_source_excel = existing_payload.get("summary", {}).get("ie_source_excel", "existing vercel data")
    else:
        raise FileNotFoundError(excel_path)

    payload = {
        "summary": {
            **ip_data.get("summary", {}),
            "service_types": ["IP", "IE"],
            "ie_fixed_rate_rows": 0,
            "ie_per_kg_rate_rows": 0,
            "ie_source_excel": ie_source_excel,
        },
        "country_alias": ip_data["country_alias"],
        "demand_surcharge_rates": ip_data.get("demand_surcharge_rates", []),
        "country_demand_region": ip_data.get("country_demand_region", []),
        "rates": {
            "IP": {
                "label": "FedEx IP 国际优先",
                "fixed_0_20_5kg": ip_data["ip_parcel_rate_0_20_5kg"],
                "perkg_21kg_plus": ip_data["ip_parcel_rate_21kg_plus"],
            },
            "IE": ie_rates,
        },
        "validation_checks": validation_checks,
    }
    payload["summary"]["ie_fixed_rate_rows"] = len(payload["rates"]["IE"]["fixed_0_20_5kg"])
    payload["summary"]["ie_per_kg_rate_rows"] = len(payload["rates"]["IE"]["perkg_21kg_plus"])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
