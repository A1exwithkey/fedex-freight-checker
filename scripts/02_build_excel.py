#!/usr/bin/env python3
"""Build the FedEx IP freight checker Excel workbook from processed JSON."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATA = PROJECT_ROOT / "data_processed" / "fedex_ip_data.json"
DEFAULT_OUTPUT = PROJECT_ROOT / "outputs" / "运费核价助手_FedEx_IP_V2.5_网页试算版.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FILL = PatternFill("solid", fgColor="17365D")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
OUTPUT_FILL = PatternFill("solid", fgColor="E2F0D9")
FINAL_FILL = PatternFill("solid", fgColor="C6E0B4")
NOTE_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_FONT = Font(name="Calibri", color="FFFFFF", bold=True)
BOLD_FONT = Font(name="Calibri", bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=16)
SMALL_FONT = Font(name="Calibri", size=9, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def add_table(ws, headers: list[str], rows: list[dict[str, Any]], widths: list[int] | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(name="Calibri", color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        width = widths[idx - 1] if widths and idx <= len(widths) else max(12, min(45, len(header) + 4))
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            if cell.row != 1:
                cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_readme(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("README")
    summary = data["summary"]
    rows = [
        {"item": "workbook_name", "value": "FedEx 运费核价助手 IP 出口 V2.5 网页试算版", "description_cn": "本文件只服务中国 FedEx 国际出口 IP 包裹核价"},
        {"item": "source_pdf", "value": summary["source_pdf"], "description_cn": "公司内部 FedEx 协议价 PDF"},
        {"item": "current_version", "value": "V2.5.0-local-web-trial-2026-05-17", "description_cn": "Excel 计算母版 + Streamlit 本地试算版数据源"},
        {"item": "scope_included", "value": "FedEx 国际优先快递服务 出口 / IP / 包裹", "description_cn": "不含快递封、快递袋、IPE、IE、进口、第三方支付、重货"},
        {"item": "rate_source_pages", "value": "PDF 7-10", "description_cn": "第 7 页下半开始为 IP；第 9-10 页为 21kg+ 每公斤费率"},
        {"item": "zone_source_pages", "value": "PDF 20-24", "description_cn": "中国 FedEx 国际出口分区列表，仅读取 IP 列"},
        {"item": "fixed_rate_rule", "value": "0.5kg-20.5kg", "description_cn": "calculator 中按 0.5kg 向上取整后查固定费率"},
        {"item": "per_kg_rule", "value": ">20.5kg", "description_cn": "calculator 中使用实际重量 × 对应重量段每公斤费率"},
        {"item": "default_fuel_surcharge", "value": 0.48, "description_cn": "可在 calculator!D4 手动修改"},
        {"item": "default_redundancy_factor", "value": 1.1, "description_cn": "冗余系数，可在 calculator!E4 手动修改"},
        {"item": "default_exchange_rate", "value": 6.8, "description_cn": "CNY/USD，可在 calculator!F4 手动修改"},
        {"item": "demand_surcharge", "value": "按目的地需求附加费区域匹配 RMB/kg；费率为 0 时不收需求附加费", "description_cn": "来源：FEDEX需求附加费-2026.4.13日生效至另行通知.pdf"},
        {"item": "demand_effective_date", "value": "2026-04-13", "description_cn": "仅使用从中国大陆出口的国际货件列"},
        {"item": "demand_minimum_rule", "value": "rate > 0 时 MAX(weight × rate, 1.80); rate = 0 时为 0", "description_cn": "PDF 第 3 页写明每票最低 1.80 元"},
        {"item": "fuel_rate_version", "value": "Manual 2026-05-17; default 48%", "description_cn": "燃油附加费每周变动，当前作为人工可编辑输入"},
        {"item": "formula", "value": "(Base Freight CNY + Demand Surcharge CNY) × (1 + Fuel Rate) × Redundancy Factor ÷ Exchange Rate", "description_cn": "所有关键公式在 calculator 中可见"},
        {"item": "data_check", "value": f"旧 Excel 费率差异：{summary['old_excel_rate_mismatch_count']}；旧国家分区换行粘连：{summary['old_excel_country_newline_zone_rows']}", "description_cn": "旧费率表可信；旧国家分区表已重建"},
    ]
    add_table(ws, ["item", "value", "description_cn"], rows, [26, 56, 66])


def add_calculator(wb: Workbook, dropdown_count: int) -> None:
    ws = wb.active
    ws.title = "calculator"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "FedEx IP 运费核价助手"
    ws.merge_cells("A1:J1")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    rows = [
        ["工具版本 2026-05-17 | IP 协议价 2026-01-05 | 需求附加费 2026-04-13 | 燃油费人工维护，默认 48%", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["输入区", "", "", "", "", "", "", "", "", ""],
        ["目的地（下拉）", "目的地（手输，可选）", "实际重量 kg", "燃油附加费率", "冗余系数", "汇率 CNY/USD", "", "", "", ""],
        ["United States - Other Areas (美国其他地区)", "", 10, 0.48, 1.1, 6.8, "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["自动匹配区", "", "", "", "", "", "", "", "", ""],
        ["实际用于匹配的目的地", "匹配国家/地区", "IP 分区", "需求附加费大区", "需求费率 CNY/kg", "最低收费 CNY/票", "查表重量", "费率类型", "状态", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["报价结果", "", "", "", "", "", "", "", "", ""],
        ["基础运费 CNY", "需求附加费 CNY", "含需求基础运费 CNY", "燃油附加费 CNY", "含冗余金额 CNY", "最终 USD", "状态", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["说明", "手输目的地优先；手输为空时使用下拉。需求附加费仅使用从中国大陆出口的国际货件列；费率为 0 时需求附加费为 0，费率大于 0 时按 MAX(重量×费率, 1.80) 计算。", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["字段名", "country_for_lookup", "matched_country", "ip_zone", "demand_region", "demand_rate_cny_per_kg", "demand_minimum_cny", "chargeable_weight_for_lookup", "rate_type", "base_freight_cny"],
        ["公式", "", "", "", "", "", "", "", "", ""],
    ]
    for row in rows:
        ws.append(row)

    match_formulas = [
        '=IF(TRIM(B5)<>"",TRIM(B5),A5)',
        '=IFERROR(VLOOKUP(LOWER(TRIM(A9)),country_alias!B:J,2,FALSE),"Need Review")',
        '=IFERROR(VLOOKUP(LOWER(TRIM(A9)),country_alias!B:J,3,FALSE),"Need Review")',
        '=IFERROR(VLOOKUP(LOWER(TRIM(A9)),country_alias!B:J,5,FALSE),"Need Review")',
        '=IFERROR(VLOOKUP(LOWER(TRIM(A9)),country_alias!B:J,7,FALSE),"Need Review")',
        '=IFERROR(VLOOKUP(LOWER(TRIM(A9)),country_alias!B:J,8,FALSE),"Need Review")',
        '=IF(C5="","",IF(C5<=20.5,CEILING(C5,0.5),C5))',
        '=IF(C5="","",IF(C5<=20.5,"固定费率 0.5-20.5kg","每公斤费率 21kg+"))',
        '=IF(OR(C9="Need Review",D9="Need Review",E9="Need Review"),"Need Review","OK")',
    ]
    result_formulas = [
        '=IF(OR(C9="Need Review",C5=""),"Need Review",IF(C5<=20.5,IF(SUMIFS(ip_parcel_rate_0_20_5kg!C:C,ip_parcel_rate_0_20_5kg!A:A,G9,ip_parcel_rate_0_20_5kg!B:B,C9)=0,"Need Review",SUMIFS(ip_parcel_rate_0_20_5kg!C:C,ip_parcel_rate_0_20_5kg!A:A,G9,ip_parcel_rate_0_20_5kg!B:B,C9)),IF(SUMIFS(ip_parcel_rate_21kg_plus!D:D,ip_parcel_rate_21kg_plus!A:A,"<="&C5,ip_parcel_rate_21kg_plus!B:B,">="&C5,ip_parcel_rate_21kg_plus!C:C,C9)=0,"Need Review",C5*SUMIFS(ip_parcel_rate_21kg_plus!D:D,ip_parcel_rate_21kg_plus!A:A,"<="&C5,ip_parcel_rate_21kg_plus!B:B,">="&C5,ip_parcel_rate_21kg_plus!C:C,C9))))',
        '=IF(ISNUMBER(E9),IF(E9>0,MAX(C5*E9,F9),0),"Need Review")',
        '=IF(AND(ISNUMBER(A13),ISNUMBER(B13)),A13+B13,"Need Review")',
        '=IF(ISNUMBER(C13),C13*D5,"Need Review")',
        '=IF(ISNUMBER(C13),((C13+D13)*E5),"Need Review")',
        '=IF(ISNUMBER(E13),E13/F5,"Need Review")',
        '=IF(OR(I9="Need Review",A13="Need Review"),"Need Review","OK")',
    ]
    for col_idx, formula in enumerate(match_formulas, start=1):
        ws.cell(row=9, column=col_idx).value = formula
        ws.cell(row=18, column=col_idx).value = "'" + formula
    for col_idx, formula in enumerate(result_formulas, start=1):
        ws.cell(row=13, column=col_idx).value = formula
        ws.cell(row=18, column=col_idx + 9).value = "'" + formula

    for row in [3, 7, 11]:
        for cell in ws[row]:
            cell.fill = TITLE_FILL
            cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A3"] = "输入区"
    ws.merge_cells("A3:J3")
    ws["A7"] = "自动匹配区"
    ws.merge_cells("A7:J7")
    ws["A11"] = "报价结果"
    ws.merge_cells("A11:J11")
    ws["A2"].fill = NOTE_FILL
    ws["A2"].font = SMALL_FONT
    ws.merge_cells("A2:J2")
    for cell in ws[4]:
        cell.fill = INPUT_FILL
    for cell in ws[5]:
        cell.fill = INPUT_FILL
    for cell in ws[8]:
        cell.fill = TITLE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[9]:
        cell.fill = OUTPUT_FILL
    for cell in ws[12]:
        cell.fill = TITLE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[13]:
        cell.fill = OUTPUT_FILL
    ws["F13"].fill = FINAL_FILL
    ws["F13"].font = Font(name="Calibri", bold=True, size=16, color="375623")
    ws["A15"].fill = NOTE_FILL
    ws.merge_cells("B15:J15")
    ws["B15"].font = SMALL_FONT
    ws["A17"].fill = NOTE_FILL
    ws["A18"].fill = NOTE_FILL

    for cell_ref in ["C5", "E5", "F5", "E9", "F9", "G9", "A13", "B13", "C13", "D13", "E13", "F13"]:
        ws[cell_ref].number_format = "0.00"
    ws["D5"].number_format = "0.00%"

    widths = [34, 30, 14, 16, 14, 15, 16, 22, 16, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            if not cell.font.bold and cell.row != 1:
                cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 34
    ws.row_dimensions[5].height = 34
    ws.row_dimensions[8].height = 34
    ws.row_dimensions[9].height = 42
    ws.row_dimensions[12].height = 34
    ws.row_dimensions[13].height = 42
    ws.freeze_panes = "A8"
    dv = DataValidation(type="list", formula1="=country_dropdown_list", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["A5"])
    ws.data_validations.dataValidation[-1].showErrorMessage = True
    ws.data_validations.dataValidation[-1].errorTitle = "请选择目的地"
    ws.data_validations.dataValidation[-1].error = "请从下拉列表选择，或在下一行手动输入。"
    ws.row_dimensions[17].hidden = True
    ws.row_dimensions[18].hidden = True


def country_excel_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        display_cn = row.get("country_region_zh") or row.get("country_region_combined") or row.get("country_region_en")
        output.append(
            {
                "display_name_cn": display_cn,
                "country_region_en": row.get("country_region_en", ""),
                "country_region_zh": row.get("country_region_zh", ""),
                "ip_zone": row.get("ip_zone", ""),
                "source_pdf_page": row.get("source_pdf_page", ""),
                "review_status": row.get("review_status", ""),
                "source_note": row.get("source_note", ""),
            }
        )
    return output


def dropdown_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    values = []
    seen = set()
    all_values = []
    for row in rows:
        en = row.get("country_region_en", "")
        zh = row.get("country_region_zh", "")
        combined = row.get("country_region_combined", "")
        if re.search(r"\d{5}", en) or re.search(r"\d{5}", zh):
            continue
        if zh == "美国其他地区":
            en = "United States - Other Areas"
        elif zh == "美国西部":
            en = "United States - Western Region"
        elif not en:
            en = combined or zh
        label = f"{en} ({zh})" if zh else en
        if label and label not in seen:
            seen.add(label)
            values.append(
                {
                    "dropdown_label": label,
                    "lookup_alias": zh or en,
                    "ip_zone": row.get("ip_zone", ""),
                    "country_region_en": en,
                    "country_region_zh": zh,
                }
            )
    values.sort(key=lambda row: row["country_region_en"].lower())
    return values


def build_workbook(data: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    wb.properties.title = "FedEx IP 运费核价助手"
    wb.properties.subject = "China FedEx IP export parcel freight checker"
    dropdown = dropdown_rows(data["country_zone_ip"])
    add_calculator(wb, len(dropdown))
    add_readme(wb, data)
    add_table(
        wb.create_sheet("country_zone_ip"),
        ["display_name_cn", "country_region_en", "country_region_zh", "ip_zone", "source_pdf_page", "review_status", "source_note"],
        country_excel_rows(data["country_zone_ip"]),
        [30, 34, 28, 10, 16, 16, 72],
    )
    dropdown_ws = wb.create_sheet("country_dropdown")
    add_table(dropdown_ws, ["dropdown_label", "lookup_alias", "ip_zone", "country_region_en", "country_region_zh"], dropdown, [44, 30, 10, 34, 28])
    dropdown_ws.sheet_state = "hidden"
    wb.defined_names.add(
        DefinedName("country_dropdown_list", attr_text=f"'country_dropdown'!$A$2:$A${len(dropdown) + 1}")
    )
    add_table(
        wb.create_sheet("ip_parcel_rate_0_20_5kg"),
        ["weight_kg", "zone", "base_rate_cny", "source_pdf_pages", "service"],
        data["ip_parcel_rate_0_20_5kg"],
        [14, 10, 18, 18, 24],
    )
    add_table(
        wb.create_sheet("ip_parcel_rate_21kg_plus"),
        ["min_kg", "max_kg", "zone", "rate_cny_per_kg", "source_pdf_pages", "service"],
        data["ip_parcel_rate_21kg_plus"],
        [12, 12, 10, 20, 18, 24],
    )
    add_table(
        wb.create_sheet("country_alias"),
        [
            "alias",
            "alias_normalized",
            "canonical_country_region",
            "ip_zone",
            "match_note",
            "demand_region_cn",
            "demand_region_code",
            "demand_rate_cny_per_kg",
            "demand_minimum_cny",
            "demand_review_status",
        ],
        data["country_alias"],
        [34, 34, 52, 10, 66, 24, 20, 20, 18, 18],
    )
    add_table(
        wb.create_sheet("demand_surcharge_rates"),
        [
            "demand_region_code",
            "demand_region_cn",
            "demand_region_en",
            "priority_rate_cny_per_kg",
            "minimum_cny_per_shipment",
            "source",
            "source_url",
            "effective_date",
            "notes",
        ],
        data["demand_surcharge_rates"],
        [20, 26, 38, 22, 24, 46, 30, 18, 82],
    )
    add_table(
        wb.create_sheet("country_demand_region"),
        [
            "country_region_cn",
            "country_region_en",
            "demand_region_code",
            "demand_region_cn",
            "demand_region_en",
            "priority_rate_cny_per_kg",
            "minimum_cny_per_shipment",
            "source",
            "effective_date",
            "review_status",
            "notes",
        ],
        data["country_demand_region"],
        [30, 34, 20, 26, 38, 22, 24, 46, 18, 16, 66],
    )
    add_table(
        wb.create_sheet("validation_checks"),
        ["test_case_id", "country_input", "matched_country", "ip_zone", "weight_kg", "pdf_page", "pdf_value", "excel_value", "pass_fail", "notes"],
        data["validation_checks"],
        [14, 34, 32, 10, 12, 12, 34, 34, 16, 62],
    )
    wb.calculation.fullCalcOnLoad = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", type=Path, default=DEFAULT_DATA)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    data = json.loads(args.data.read_text(encoding="utf-8"))
    build_workbook(data, args.output)
    print(args.output)


if __name__ == "__main__":
    main()
